"""Domus budget app — Streamlit. Layout and levy math follow Domus Excel packs."""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Domus Budget", layout="wide", initial_sidebar_state="expanded")

YELLOW, NAVY, BLUE, RED, TOTAL, SECTION = "FFFF99", "1F4E79", "0000FF", "FFC7CE", "D9E2F3", "2E75B6"
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
MONEY = '#,##0.00;(#,##0.00);"-"'


def uid() -> str:
    import random, string
    return "".join(random.choices(string.ascii_lowercase + string.digits, k=8))


def money(n: float) -> str:
    return f"R {n:,.2f}"


def norm(s: str) -> str:
    s = str(s or "").lower()
    s = re.sub(r"[–—-]", " ", s)
    s = re.sub(r"[^a-z0-9/& ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def row(desc: str, note: str = "") -> dict:
    return {
        "id": uid(), "desc": desc, "actual": 0.0, "pct": 0.0, "yearly": 0.0,
        "insurance": 0.0, "note": note, "is_recovery": False,
    }


def default_sections() -> dict:
    return {
        "levy": [
            row("Ordinary Levies", "Admin levy. Does not include estate/HOA pass-throughs or insurance billed on its own PQ column."),
            row("Reserve Fund Contribution", "Type the yearly amount, or use 15% of ordinary."),
            row("CSOS Levy (Income)", "This complex’s own CSOS. Not the master-estate CSOS."),
            row("Insurance billed to owners", "Thornhill / Mount Kos style. Own PQ column. Leave 0 if insurance stays inside ordinary."),
            row("Levy - Boathouse"),
            row("Levy - Boatport"),
            row("Special Levy"),
        ],
        "other": [
            row("Interest on Arrear Levies", "Usually not budgeted (do not rely on arrears)."),
            row("Investment Income"),
            row("Penalty Income"),
            row("Eskom / Electricity meters fixed charge", "Only if this complex bills a fixed charge."),
            row("Communal electricity recovered"),
            row("Rental Income"),
            row("Garage Rental Income"),
            row("Gate Registration / Services"),
            row("Clubhouse Rental"),
        ],
        "hoa_income": [
            row("Estate / HOA Levies recovered", "e.g. Xanadu Levies. Billed to owners on the PQ. Not ordinary."),
            row("Estate / HOA CSOS recovered", "e.g. Xanadu HOA CSOS. Own PQ column."),
        ],
        "hoa_expense": [
            row("Estate / HOA Levies paid", "e.g. Xanadu Eco Park. Paid to the estate. Not in ordinary."),
            row("Estate / HOA CSOS paid", "e.g. Xanadu CSOS paid. Not in ordinary."),
        ],
        "recoveries_other": [
            row("Insurance claims recovered", "Claim payouts. Deduct on the R&M line. Do not budget as normal income."),
            row("Legal Fees Recovered"),
            row("Maintenance Recovered"),
        ],
        "municipal": [
            row("Electricity"),
            {**row("Less: Electricity recovered from owners"), "is_recovery": True},
            row("Water"),
            {**row("Less: Water recovered from owners"), "is_recovery": True},
            row("Sewerage"),
            {**row("Less: Sewerage recovered from owners"), "is_recovery": True},
            row("Refuse Removal"),
            {**row("Less: Refuse recovered from owners"), "is_recovery": True},
            row("Rates / Property Tax"),
        ],
        "expenditure": [
            row("Accounting Fees"), row("Audit Fees"), row("Bank Charges"),
            row("CSOS Levies (Expense)"), row("Insurance"), row("Management Fees"),
            row("Legal Expense"), row("Security / Guarding"), row("Cleaning Materials"),
            row("Computer Expenses"), row("Printing and Stationery"),
            row("Telephone and Internet"), row("Meeting Expenses"),
            row("Health & Safety"), row("Protective Clothing"),
            row("Office / General Expenses"), row("Property Valuation"),
            row("Garden service (contract)"), row("Motor Vehicle Expense"),
        ],
        "rm": [
            row("Electrical"), row("Fire Equipment"), row("General Building"),
            row("Plumbing / Sewerage"), row("Gate & Intercom"), row("Garden Expenses"),
            row("Roofs & Gutters"), row("Painting / Waterproofing"), row("Pool"),
            row("Electric Fence"), row("CCTV"), row("Paving / Roadways"),
            row("Lifts"), row("Equipment Repairs"), row("Other R&M"),
        ],
        "personnel": [
            row("Salaries & Wages"), row("Casual / Relief Wages"), row("PAYE / UIF"),
            row("Travel"), row("Bonuses & Overtime"), row("WCA / COIDA"),
            row("Pension / Provident Fund"), row("Staff Welfare"), row("Caretaker Fees"),
        ],
        "tax": [row("Taxation Payable", "Based on taxable investment income. Do not leave this blank if the FS has tax.")],
        "special": [
            row("Special Project 1"), row("Special Project 2"), row("Special Project 3"),
        ],
        "fixed": [
            row("Insurance billed to owners (monthly fixed)", "Mount Kos-style extra on the levy invoice."),
            row("Prepaid meters estimate (monthly)"),
            row("Eskom fixed charge (monthly)"),
            row("Communal charge (monthly)"),
        ],
    }


def net_of(r: dict) -> float:
    y = float(r.get("yearly") or 0)
    ins = float(r.get("insurance") or 0)
    if r.get("is_recovery"):
        return -abs(y)
    return y - ins


def sum_net(items: list) -> float:
    return sum(net_of(r) for r in items)


def insurance_on_pq(state: dict) -> bool:
    for r in state["sections"].get("levy", []):
        if "insurance" in r["desc"].lower() and float(r.get("yearly") or 0) > 0.5:
            return True
    return False


def skip_from_ordinary(r: dict, state: dict) -> bool:
    f = family(r.get("desc") or "")
    if f in ("hoa_levy_exp", "hoa_csos_exp", "hoa_levy_inc", "hoa_csos_inc"):
        return True
    if f == "insurance" and insurance_on_pq(state):
        return True
    return False


def ordinary_total(state: dict) -> float:
    s = state["sections"]
    total = sum_net(s["municipal"]) + sum_net(s["rm"]) + sum_net(s["personnel"]) + sum_net(s["tax"])
    total += sum(net_of(r) for r in s["expenditure"] if not skip_from_ordinary(r, state))
    if state.get("special_in_ordinary"):
        total += sum_net(s["special"])
    return total


def apply_levy_lines(state: dict) -> None:
    ord_amt = ordinary_total(state)
    for r in state["sections"]["levy"]:
        d = r["desc"].lower()
        if "ordinary" in d:
            r["yearly"] = ord_amt
            a = float(r.get("actual") or 0)
            r["pct"] = 0.0 if a == 0 else (ord_amt / a) * 100 - 100
        if "reserve" in d:
            if state.get("reserve_mode") == "15pct":
                r["yearly"] = ord_amt * 0.15
            else:
                r["yearly"] = float(state.get("reserve_amount") or r.get("yearly") or 0)
            a = float(r.get("actual") or 0)
            r["pct"] = 0.0 if a == 0 else (float(r["yearly"]) / a) * 100 - 100


def pq_bill_lines(state: dict) -> list:
    """Owner-invoice columns, same order as Thornhill PQ."""
    s = state["sections"]
    out = []

    def add(name, yearly):
        out.append((str(name), float(yearly or 0)))

    for r in s.get("levy", []):
        if "ordinary" in r["desc"].lower() or r["desc"].lower().strip() in ("levies", "levy"):
            add("Levies", r.get("yearly"))
            break
    for r in s.get("hoa_income", []):
        if float(r.get("yearly") or 0) or float(r.get("actual") or 0):
            add(r["desc"], r.get("yearly"))
    for r in s.get("levy", []):
        if "insurance" in r["desc"].lower() and (float(r.get("yearly") or 0) or float(r.get("actual") or 0)):
            add("Insurance", r.get("yearly"))
    for r in s.get("levy", []):
        if "reserve" in r["desc"].lower():
            add("Reserve Fund", r.get("yearly"))
            break
    for r in s.get("levy", []):
        if "csos" in r["desc"].lower():
            add("CSOS", r.get("yearly"))
            break
    for r in s.get("levy", []):
        d = r["desc"].lower()
        if any(x in d for x in ("boathouse", "boatport", "special levy")) and (
            float(r.get("yearly") or 0) or float(r.get("actual") or 0)
        ):
            add(r["desc"], r.get("yearly"))
    return out


def family(desc: str) -> str | None:
    d = norm(desc)
    master = bool(re.search(r"xanadu|eco park|master scheme|master hoa|\bhoa\b|estate levy|estate levies", d))
    if master:
        if "csos" in d:
            if re.search(r"hoa csos|csos recovered|csos income", d):
                return "hoa_csos_inc"
            if re.search(r"xanadu csos|csos paid|csos expense", d) and "hoa" not in d:
                return "hoa_csos_exp"
            return "hoa_csos_inc"
        if "eco park" in d or re.search(r"paid|expense", d):
            return "hoa_levy_exp"
        return "hoa_levy_inc"
    if re.search(r"insurance\s*(claim|payout)", d):
        return "ins_claim"
    if re.search(r"insurance\s*(recovered|billed|additional)", d):
        return "ins_bill"
    if "eskom" in d or ("fixed" in d and "electr" in d) or "meters recovered" in d:
        return "eskom"
    if "reserve" in d:
        return "reserve"
    if "special levy" in d:
        return "special_levy"
    if "csos" in d:
        if "collect" in d:
            return "csos_col"
        if "contrib" in d or "expense" in d or "admin" in d:
            return "csos_exp"
        return "csos_inc"
    if "boathouse" in d:
        return "boathouse"
    if "boatport" in d:
        return "boatport"
    if d in ("levies", "levy") or "ordinary" in d or re.search(r"^levies?\s*(unit|received)?$", d):
        return "ordinary"
    if "electricity" in d and "recover" in d and "commun" in d:
        return "elec_comm"
    if "electricity" in d and "recover" in d:
        return "elec_rec"
    if "water" in d and "recover" in d:
        return "water_rec"
    if "sewer" in d and "recover" in d:
        return "sewer_rec"
    if "refuse" in d and "recover" in d:
        return "refuse_rec"
    if "interest" in d and "arrear" in d:
        return "int_arr"
    if re.search(r"interest|investment|marketlink", d) and re.search(r"bank|invest|marketlink", d):
        return "invest"
    if "penalty" in d:
        return "penalty"
    if "rental" in d and "garage" in d:
        return "garage"
    if "rental" in d or "rent received" in d:
        return "rental"
    if "electricity" in d and "recover" not in d:
        return "elec_g"
    if re.search(r"^water$|water\s*(charge|expense)", d):
        return "water_g"
    if "sewer" in d and "recover" not in d:
        return "sewer_g"
    if "refuse" in d and "recover" not in d:
        return "refuse_g"
    if "management" in d and "fee" in d:
        return "mgmt"
    if re.search(r"^insurance$|insurance\s*premium", d):
        return "insurance"
    if "security" in d or "guarding" in d:
        return "security"
    if re.search(r"salar|staff wages", d):
        return "salaries"
    return None


def section_for(desc: str) -> str:
    f = family(desc)
    d = norm(desc)
    if f in ("ordinary", "reserve", "csos_inc", "boathouse", "boatport", "ins_bill", "special_levy"):
        return "levy"
    if f in ("hoa_levy_inc", "hoa_csos_inc"):
        return "hoa_income"
    if f in ("hoa_levy_exp", "hoa_csos_exp"):
        return "hoa_expense"
    if f == "csos_exp":
        return "expenditure"
    if f == "csos_col":
        return "other"
    if f in ("elec_rec", "water_rec", "sewer_rec", "refuse_rec"):
        return "municipal"
    if f in ("elec_g", "water_g", "sewer_g", "refuse_g"):
        return "municipal"
    if f in ("eskom", "elec_comm", "int_arr", "invest", "penalty", "rental", "garage"):
        return "other"
    if f == "ins_claim":
        return "recoveries_other"
    if f == "salaries":
        return "personnel"
    if re.search(r"\b(repair|maintenance|plumb|paint|roof|gutter|pool|electrical|fire equipment|gate|paving)\b", d):
        return "rm"
    if re.search(r"\b(wages|salary|paye|uif|bonus|overtime|casual|relief|wca|coida|caretaker|staff)\b", d):
        return "personnel"
    if re.search(r"\b(income tax|taxation)\b", d):
        return "tax"
    if re.search(r"\b(special project|improvement|jungle gym|damp|aluminium)\b", d):
        return "special"
    if "legal" in d and "recover" in d:
        return "recoveries_other"
    if "recover" in d:
        return "recoveries_other"
    return "expenditure"


def num(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        n = float(v)
        return 0.0 if abs(n) < 0.01 else n
    s = str(v).strip()
    if not s or s in ("-", "–"):
        return 0.0
    s2 = re.sub(r"[R$\s,]", "", s).replace("(", "-").replace(")", "")
    try:
        n = float(s2)
    except ValueError:
        return None
    return 0.0 if abs(n) < 0.01 else n


def extract_wcu(uploaded) -> list:
    xl = pd.ExcelFile(uploaded)
    out = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, header=None)
        if df.empty or df.shape[1] < 4:
            continue
        header = None
        for i in range(min(15, len(df))):
            vals = [str(v).strip().lower() for v in df.iloc[i].tolist()]
            if "actual" in vals and any("budget" in v for v in vals):
                header = i
                break
        if header is None:
            continue
        prev = [str(v).strip().lower() for v in df.iloc[header - 1].tolist()] if header else [""] * df.shape[1]
        cur = [str(v).strip().lower() for v in df.iloc[header].tolist()]
        merged = [(prev[i] if i < len(prev) else "") + " " + (cur[i] if i < len(cur) else "") for i in range(df.shape[1])]
        ytd = next((i for i, t in enumerate(merged) if "ytd" in t and "actual" in t and "var" not in t), None)
        if ytd is None:
            ytd = next((i for i, t in enumerate(cur) if t == "actual"), None)
        if ytd is None:
            continue
        for r in range(header + 1, len(df)):
            raw = str(df.iloc[r, 0] or "").strip()
            if not raw or re.match(r"^(total|surplus|shortfall)", raw, re.I):
                continue
            m = re.match(r"^(\d{3,5}\s*/\s*\d{2,4})\s*[-–—:]\s*(.+)$", raw)
            if m:
                gl = re.sub(r"\s", "", m.group(1))
                desc = m.group(2).strip()
                if gl.endswith("/000"):
                    continue
            else:
                desc = raw
                if len(desc) < 3:
                    continue
            actual = abs(num(df.iloc[r, ytd]) or 0.0)
            if actual < 0.5:
                continue
            out.append({"desc": desc, "actual": actual})
    return out


def match_into(extracted: list, sections: dict) -> tuple[dict, int]:
    nxt = {k: [dict(x) for x in v] for k, v in sections.items()}
    used = set()
    added = 0
    flat = [(k, it) for k, items in nxt.items() for it in items]
    leftover = []
    for src in extracted:
        fam = family(src["desc"])
        if fam == "csos_col":
            continue
        best, score = None, 0.0
        for key, it in flat:
            mark = f"{key}:{it['id']}"
            if mark in used:
                continue
            if fam and family(it["desc"]) and fam != family(it["desc"]):
                continue
            if ("reserve" in norm(src["desc"]) and "eskom" in norm(it["desc"])) or (
                "eskom" in norm(src["desc"]) and "reserve" in norm(it["desc"])
            ):
                continue
            nd, ni = norm(src["desc"]), norm(it["desc"])
            sc = 0.0
            if fam and family(it["desc"]) == fam:
                sc = 0.96
            if nd == ni:
                sc = 1.0
            elif ni in nd or nd in ni:
                sc = max(sc, 0.86)
            else:
                stop = {"levy", "levies", "income", "other", "general", "expense", "expenses", "fee", "fees", "and", "the"}
                a = {w for w in nd.split() if len(w) > 3 and w not in stop}
                b = {w for w in ni.split() if len(w) > 3 and w not in stop}
                if a and b:
                    sc = max(sc, len(a & b) / max(len(a), len(b)))
            if sc > 0.55 and sc > score:
                score, best = sc, (key, it)
        if best:
            key, it = best
            used.add(f"{key}:{it['id']}")
            it["actual"] = src["actual"]
            if fam in ("hoa_levy_inc", "hoa_csos_inc", "hoa_levy_exp", "hoa_csos_exp", "ins_bill"):
                it["desc"] = src["desc"]
            if it.get("is_recovery"):
                it["yearly"] = src["actual"]
            else:
                it["yearly"] = src["actual"] * (1 + float(it.get("pct") or 0) / 100)
        else:
            leftover.append(src)
    for src in leftover:
        fam = family(src["desc"])
        if fam == "csos_col":
            continue
        sec = section_for(src["desc"])
        if fam:
            existing = next((i for i in nxt[sec] if family(i["desc"]) == fam), None)
            if existing:
                existing["actual"] = float(existing["actual"] or 0) + src["actual"]
                continue
        extra = row(src["desc"], "Added from WeConnectU for this complex")
        extra["actual"] = src["actual"]
        extra["yearly"] = src["actual"]
        extra["is_recovery"] = "recover" in norm(src["desc"]) and sec == "municipal"
        nxt[sec].append(extra)
        added += 1
    return nxt, added


def items_to_df(items: list, rm: bool) -> pd.DataFrame:
    recs = []
    for it in items:
        recs.append({
            "Description": it["desc"],
            "Actual": float(it.get("actual") or 0),
            "% Increase": float(it.get("pct") or 0),
            "Budgeted yearly": float(it.get("yearly") or 0),
            "Monthly": net_of(it) / 12,
            "Insurance payout": float(it.get("insurance") or 0),
            "Notes": it.get("note") or "",
        })
    cols = ["Description", "Actual", "% Increase", "Budgeted yearly", "Monthly", "Notes"]
    if rm:
        cols = ["Description", "Actual", "% Increase", "Budgeted yearly", "Monthly", "Insurance payout", "Notes"]
    return pd.DataFrame(recs)[cols]


def save_editor(edited: pd.DataFrame, previous: list, rm: bool) -> list:
    out = []
    records = edited.to_dict("records")
    for i, rec in enumerate(records):
        desc = str(rec.get("Description") or "").strip()
        if not desc:
            continue
        prev = previous[i] if i < len(previous) else {}
        actual = float(rec.get("Actual") or 0)
        pct = float(rec.get("% Increase") or 0)
        yearly = float(rec.get("Budgeted yearly") or 0)
        ins = float(rec.get("Insurance payout") or 0) if rm else float(prev.get("insurance") or 0)
        old_pct = float(prev.get("pct") or 0)
        old_y = float(prev.get("yearly") or 0)
        pct_changed = abs(pct - old_pct) > 0.05
        y_changed = abs(yearly - old_y) > 0.02
        if pct_changed and not y_changed:
            yearly = actual * (1 + pct / 100)
        elif y_changed:
            pct = 0.0 if actual == 0 else (yearly / actual) * 100 - 100
        out.append({
            "id": prev.get("id") or uid(),
            "desc": desc,
            "actual": actual,
            "pct": pct,
            "yearly": yearly,
            "insurance": ins,
            "note": str(rec.get("Notes") or ""),
            "is_recovery": bool(prev.get("is_recovery")) or desc.lower().startswith("less:"),
        })
    return out


def generate_excel(state: dict) -> BytesIO:
    apply_levy_lines(state)
    s = state["sections"]
    wb = Workbook()
    ws = wb.active
    ws.title = "BUDGET"
    for i, w in enumerate([3, 44, 16, 12, 16, 14, 14, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def fill(c, color):
        c.fill = PatternFill("solid", fgColor=color)

    def inp(c, val, fmt=None):
        c.value = val
        fill(c, YELLOW)
        c.font = Font(name="Calibri", color=BLUE, size=10)
        c.border = THIN
        if fmt:
            c.number_format = fmt

    def fml(c, f, bg=None):
        c.value = f"={f}"
        c.font = Font(name="Calibri", size=10)
        c.border = THIN
        c.number_format = MONEY
        if bg:
            fill(c, bg)

    ws.merge_cells("B2:G2")
    ws["B2"] = "BODY CORPORATE / HOA BUDGET"
    ws["B2"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    ws["B3"] = "Complex:"
    inp(ws["C3"], state.get("complex_name") or "")
    ws["E3"] = "Year:"
    inp(ws["F3"], state.get("fin_year") or "")
    ws["B5"] = "Ordinary levies = net municipal + expenditure + R&M (after insurance) + personnel + tax"
    ws["B6"] = "Reserve method:"
    inp(ws["C6"], "15% of ordinary" if state.get("reserve_mode") == "15pct" else "Typed amount")

    r = 8

    def bar(title):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        for col in range(2, 8):
            fill(ws.cell(r, col), SECTION)
            ws.cell(r, col).border = THIN
        ws.cell(r, 2).value = title
        ws.cell(r, 2).font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        r += 1

    def hdr(rm=False):
        nonlocal r
        labs = ["Description", "Actual", "% Increase", "Budgeted Yearly", "Monthly", "Notes"]
        if rm:
            labs = ["Description", "Actual", "% Increase", "Budgeted Yearly", "Insurance payout", "Monthly", "Notes"]
        for i, lab in enumerate(labs, 2):
            c = ws.cell(r, i, lab)
            fill(c, NAVY)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.border = THIN
        r += 1

    def write(items, rm=False):
        nonlocal r
        start = r
        for it in items:
            inp(ws.cell(r, 2), it["desc"])
            inp(ws.cell(r, 3), float(it.get("actual") or 0), MONEY)
            inp(ws.cell(r, 4), float(it.get("pct") or 0) / 100, "0.0%")
            y = float(it.get("yearly") or 0)
            if it.get("is_recovery"):
                inp(ws.cell(r, 5), -abs(y), MONEY)
            else:
                inp(ws.cell(r, 5), y, MONEY)
            if rm:
                inp(ws.cell(r, 6), float(it.get("insurance") or 0), MONEY)
                fml(ws.cell(r, 7), f"(E{r}-F{r})/12")
                inp(ws.cell(r, 8), it.get("note") or "")
            else:
                fml(ws.cell(r, 6), f"E{r}/12")
                inp(ws.cell(r, 7), it.get("note") or "")
            r += 1
        return start, r - 1

    def tot(label, start, end, rm=False):
        nonlocal r
        ws.cell(r, 2, label).font = Font(bold=True)
        if rm:
            fml(ws.cell(r, 5), f"SUM(E{start}:E{end})-SUM(F{start}:F{end})", TOTAL)
            fml(ws.cell(r, 7), f"E{r}/12", TOTAL)
        else:
            fml(ws.cell(r, 5), f"SUM(E{start}:E{end})", TOTAL)
            fml(ws.cell(r, 6), f"E{r}/12", TOTAL)
        r += 2

    bar("INCOME — LEVY INCOME")
    hdr()
    a, b = write(s["levy"])
    levy_ord = a
    tot("TOTAL LEVY INCOME", a, b)
    bar("OTHER INCOME")
    hdr()
    a, b = write(s["other"])
    tot("TOTAL OTHER INCOME", a, b)
    bar("OTHER RECOVERIES (not municipal utilities)")
    hdr()
    a, b = write(s["recoveries_other"])
    tot("TOTAL OTHER RECOVERIES", a, b)
    bar("HOA / ESTATE RECOVERED FROM OWNERS (optional — Thornhill / Xanadu)")
    hdr()
    a, b = write(s.get("hoa_income") or [])
    tot("TOTAL HOA RECOVERED", a, b)
    bar("HOA / ESTATE PAID TO THE ESTATE (optional)")
    hdr()
    a, b = write(s.get("hoa_expense") or [])
    tot("TOTAL HOA PAID", a, b)
    hdr()
    a, b = write(s["municipal"])
    muni_tot = r
    tot("NET MUNICIPAL CHARGES", a, b)
    bar("EXPENDITURE")
    hdr()
    a, b = write(s["expenditure"])
    exp_tot = r
    tot("TOTAL EXPENDITURE", a, b)
    bar("REPAIR AND MAINTENANCE")
    hdr(rm=True)
    a, b = write(s["rm"], rm=True)
    rm_tot = r
    tot("NET R&M (after insurance payouts)", a, b, rm=True)
    bar("PERSONNEL")
    hdr()
    a, b = write(s["personnel"])
    per_tot = r
    tot("TOTAL PERSONNEL", a, b)
    bar("INCOME TAX")
    hdr()
    a, b = write(s["tax"])
    tax_tot = r
    tot("TOTAL TAX", a, b)
    bar("SPECIAL PROJECTS")
    hdr()
    a, b = write(s["special"])
    tot("TOTAL SPECIAL PROJECTS", a, b)
    bar("ORDINARY LEVY CHECK")
    ws.cell(r, 2, "Ordinary levies (must equal net municipal + expenditure + net R&M + personnel + tax)")
    fml(ws.cell(r, 5), f"E{muni_tot}+E{exp_tot}+E{rm_tot}+E{per_tot}+E{tax_tot}", RED)
    ws.cell(levy_ord, 5).value = f"=E{r}"
    r += 3
    bar("FIXED MONTHLY CHARGES ON THE OWNER INVOICE (optional)")
    hdr()
    write(s["fixed"])

    pq = wb.create_sheet("PQ")
    pq["A1"] = "PQ / LEVY SCHEDULE"
    pq["A1"].font = Font(bold=True, size=14, color=NAVY)
    bills = pq_bill_lines(state)
    pq["A3"] = "Monthly totals billed to all owners"
    for i, (name, yearly) in enumerate(bills):
        pq.cell(4, i + 1, name)
        fill(pq.cell(4, i + 1), NAVY)
        pq.cell(4, i + 1).font = Font(bold=True, color="FFFFFF")
        pq.cell(5, i + 1, yearly / 12).number_format = MONEY
        fill(pq.cell(5, i + 1), YELLOW)
    headers = ["#", "Unit", "PQ"] + [n for n, _ in bills] + ["Total"]
    for i, h in enumerate(headers, 1):
        cell = pq.cell(8, i, h)
        fill(cell, NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
    units = state.get("pq") or [{"Unit": "UNIT-1", "PQ": 1.0}]
    first_amt = 4
    last_amt = 3 + len(bills)
    for i, u in enumerate(units):
        rr = 9 + i
        pq.cell(rr, 1, i + 1)
        pq.cell(rr, 2, str(u.get("Unit", "")))
        pq.cell(rr, 3, float(u.get("PQ") or 0)).number_format = "0.000000"
        for j in range(len(bills)):
            col = 4 + j
            letter = get_column_letter(j + 1)
            pq.cell(rr, col, f"=$C{rr}*{letter}$5").number_format = MONEY
        if bills:
            pq.cell(rr, last_amt + 1, f"=SUM({get_column_letter(first_amt)}{rr}:{get_column_letter(last_amt)}{rr})").number_format = MONEY

    ymp = wb.create_sheet("10 YMP")
    ymp["A1"] = "10 YEAR MAINTENANCE PLAN"
    ymp["A1"].font = Font(bold=True, size=14, color=NAVY)
    ymp.cell(3, 1, "Project")
    fill(ymp.cell(3, 1), NAVY)
    ymp.cell(3, 1).font = Font(bold=True, color="FFFFFF")
    for y in range(10):
        c = ymp.cell(3, 2 + y, f"Year {y+1}")
        fill(c, NAVY)
        c.font = Font(bold=True, color="FFFFFF")
    ymp.column_dimensions["A"].width = 36
    projects = state.get("ymp") or [{"desc": "Project 1", "years": [0] * 10}]
    for i, p in enumerate(projects):
        rr = 4 + i
        inp(ymp.cell(rr, 1), p.get("desc") or "")
        years = p.get("years") or [0] * 10
        for y in range(10):
            inp(ymp.cell(rr, 2 + y), float(years[y] if y < len(years) else 0), MONEY)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def init():
    ss = st.session_state
    ss.setdefault("sections", default_sections())
    ss.setdefault("complex_name", "")
    ss.setdefault("fin_year", "01-03-2026 / 28-02-2027")
    ss.setdefault("reserve_mode", "amount")
    ss.setdefault("reserve_amount", 0.0)
    ss.setdefault("special_in_ordinary", False)
    ss.setdefault("pq", None)
    ss.setdefault("ymp", [{"desc": "", "years": [0.0] * 10}])
    ss.setdefault("msg", "")
    ss.setdefault("current_monthly_levy", 0.0)
    ss.setdefault("actual_months", 12)
    ss.setdefault("estate_levy_yearly", 0.0)
    ss.setdefault("estate_levy_name", "Estate / master HOA levy")
    ss.setdefault("estate_levy_mode", "separate")
    ss.setdefault("estate_split", "equal")


def section_form(key: str, title: str, help_text: str, rm: bool = False):
    st.subheader(title)
    if help_text:
        st.caption(help_text)
    items = st.session_state.sections[key]
    df = items_to_df(items, rm)
    with st.form(f"form_{key}"):
        edited = st.data_editor(
            df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Description": st.column_config.TextColumn("Description", width="medium"),
                "Actual": st.column_config.NumberColumn("Actual", format="%.2f"),
                "% Increase": st.column_config.NumberColumn("% Increase", format="%.1f", help="Type % then click Save"),
                "Budgeted yearly": st.column_config.NumberColumn("Budgeted yearly", format="%.2f", help="Or type the rand amount then Save"),
                "Monthly": st.column_config.NumberColumn("Monthly", format="%.2f", disabled=True, help="Yearly ÷ 12. Updates when you Save."),
                "Insurance payout": st.column_config.NumberColumn("Insurance payout", format="%.2f"),
                "Notes": st.column_config.TextColumn("Notes"),
            },
            disabled=["Monthly"],
        )
        saved = st.form_submit_button("Save this section", type="primary")
    if saved:
        st.session_state.sections[key] = save_editor(edited, items, rm)
        apply_levy_lines(st.session_state)
        st.success("Saved. Monthly = yearly ÷ 12. % = (yearly ÷ actual) × 100 − 100.")
        st.rerun()
    st.caption(f"Section net total: {money(sum_net(st.session_state.sections[key]))}  ·  Monthly total: {money(sum_net(st.session_state.sections[key]) / 12)}")


def main():
    init()
    logo = Path(__file__).parent / "domus_logo.jpeg"
    cols = st.columns([1, 5])
    with cols[0]:
        if logo.exists():
            st.image(str(logo), width=110)
    with cols[1]:
        st.title("Domus Property Management Budget")
        st.caption("Load WeConnectU → change a section → click Save once → download Excel")

    apply_levy_lines(st.session_state)
    s = st.session_state.sections
    ord_amt = ordinary_total(st.session_state)
    reserve = next((r for r in s["levy"] if "reserve" in r["desc"].lower()), None)
    ordinary_row = next((r for r in s["levy"] if "ordinary" in r["desc"].lower()), None)
    actual_year = float(ordinary_row["actual"]) if ordinary_row else 0.0
    months = max(1, int(st.session_state.get("actual_months") or 12))
    if months < 12 and actual_year > 0:
        actual_year_full = actual_year / months * 12
    else:
        actual_year_full = actual_year
    current_m = float(st.session_state.get("current_monthly_levy") or 0)
    if current_m <= 0 and actual_year_full > 0:
        current_m = actual_year_full / 12
    new_m = ord_amt / 12
    levy_pct = 0.0 if current_m == 0 else (new_m / current_m) * 100 - 100

    if actual_year > 0 and ord_amt > 0 and 8 <= (ord_amt / actual_year) <= 15:
        st.error(
            f"Ordinary **Actual** looks like one month ({money(actual_year)}), but "
            f"**Budgeted yearly** is a full year ({money(ord_amt)}). "
            f"Put the full-year levy in Actual (about {money(actual_year * 12)}), "
            f"or type the current monthly total in the sidebar. "
            f"Trustees should compare {money(actual_year)} / month with {money(new_m)} / month."
        )

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("What owners pay now (monthly)", money(current_m))
    m2.metric("What owners will pay (monthly)", money(new_m), delta=f"{levy_pct:+.1f}%")
    m3.metric("Ordinary levies for the year", money(ord_amt))
    m4.metric("Reserve for the year", money(float(reserve["yearly"]) if reserve else 0))

    with st.sidebar:
        st.header("Complex")
        st.session_state.complex_name = st.text_input("Complex name", st.session_state.complex_name)
        st.session_state.fin_year = st.text_input("Financial year", st.session_state.fin_year)

        st.header("Load last year")
        up = st.file_uploader("WeConnectU Actual vs Budget (Excel)", type=["xlsx", "xls", "xlsm"])
        if up and st.button("Load Excel", type="primary"):
            try:
                rows = extract_wcu(up)
                if not rows:
                    st.error("No lines found. Export Options → Budget and Actuals.")
                else:
                    secs, added = match_into(rows, default_sections())
                    st.session_state.sections = secs
                    st.session_state.msg = f"Loaded {len(rows)} lines. Extra lines for this complex: {added}."
                    apply_levy_lines(st.session_state)
                    st.rerun()
            except Exception as e:
                st.error(f"Could not read file: {e}")

        st.header("What owners pay now")
        st.session_state.current_monthly_levy = st.number_input(
            "Current ordinary levy — all units, one month",
            value=float(st.session_state.current_monthly_levy),
            min_value=0.0,
            step=100.0,
            help="Example: R26 400 per month for the whole complex. Not the yearly total.",
        )
        st.session_state.actual_months = st.number_input(
            "Months covered by the Actual column",
            min_value=1,
            max_value=12,
            value=int(st.session_state.actual_months),
            help="12 = a full year. If WeConnectU is only 6 months, put 6 and we scale up for the %.",
        )
        st.header("Reserve fund")
        st.session_state.reserve_mode = st.radio(
            "How is reserve calculated?",
            ["amount", "15pct"],
            format_func=lambda x: "I will type the yearly amount" if x == "amount" else "15% of ordinary levies",
            index=0 if st.session_state.reserve_mode == "amount" else 1,
        )
        if st.session_state.reserve_mode == "amount":
            st.session_state.reserve_amount = st.number_input(
                "Reserve fund contribution (yearly rands)",
                value=float(st.session_state.reserve_amount),
                step=1000.0,
                min_value=0.0,
            )
        st.session_state.special_in_ordinary = st.checkbox(
            "Add Special Projects into ordinary levies",
            value=st.session_state.special_in_ordinary,
            help="Tick only if special work is paid from levies, not from the reserve fund.",
        )
        st.header("Second levy (master estate)")
        st.caption("Use this when owners also pay another estate / HOA.")
        st.session_state.estate_levy_name = st.text_input(
            "Name on the owner schedule",
            st.session_state.estate_levy_name,
        )
        st.session_state.estate_levy_mode = st.radio(
            "Who collects it?",
            ["separate", "we_collect"],
            format_func=lambda x: (
                "Owners pay the estate themselves — do not put it in our ordinary levy"
                if x == "separate"
                else "We collect it and pay the estate"
            ),
            index=0 if st.session_state.estate_levy_mode == "separate" else 1,
        )
        st.session_state.estate_levy_yearly = st.number_input(
            "Estate levy for the whole complex (yearly rands)",
            value=float(st.session_state.estate_levy_yearly),
            min_value=0.0,
            step=1000.0,
        )
        st.session_state.estate_split = st.radio(
            "How is it split per unit?",
            ["equal", "pq"],
            format_func=lambda x: "Same amount each unit" if x == "equal" else "By PQ",
            index=0 if st.session_state.estate_split == "equal" else 1,
        )
        if st.button("Start over"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    if st.session_state.msg:
        st.success(st.session_state.msg)

    tabs = st.tabs([
        "How it works", "Income", "Municipal", "Expenditure",
        "Repair & Maintenance", "Personnel", "Tax", "Special",
        "PQ / Levies", "10-year plan", "Download",
    ])

    with tabs[0]:
        st.markdown(
            """
### One rule for typing
Change the numbers, then click **Save this section** once. Nothing is stored until Save.
That stops the amount jumping back to 0.

- Type **% Increase** and Save → Budgeted yearly = Actual × (1 + %).
- Type **Budgeted yearly** and Save → % = (Yearly ÷ Actual) × 100 − 100.
- **Monthly** is always yearly ÷ 12.

### Ordinary levies (same as LTP / Depotel / Mount Kos)
**Ordinary = Net municipal + Expenditure + R&M after insurance + Personnel + Tax**

Net municipal means gross electricity/water/sewer/refuse minus recoveries from owners.

Reserve and CSOS are billed on their own PQ columns. They are not folded into ordinary.

### Second levy (inside another estate) — Thornhill / Xanadu
Xanadu invoices **Thornhill BC**. Thornhill bills owners on the PQ as extra columns:

- Levies (ordinary admin)
- Xanadu Levies
- Xanadu HOA CSOS
- Insurance
- Reserve Fund
- CSOS (this complex)

Those estate lines are **not** folded into ordinary. Leave them 0 on a standalone complex.

### Reserve
Type the yearly rand amount, **or** choose 15% of ordinary (Matte Court style).

### Insurance claims
On **Repair & Maintenance**, type the payout against the **specific line** (roof, plumbing, …).
That line’s net = budgeted yearly − insurance payout.
            """
        )

    with tabs[1]:
        st.info("Ordinary and Reserve update when you save the cost sections / sidebar.")
        section_form("levy", "Levy Income", "Add boathouse / boatport / extra levy types with a new row, then Save.")
        st.divider()
        section_form(
            "hoa_income",
            "Estate / HOA recovered from owners",
            "Only for complexes inside another estate (Thornhill / Xanadu). These become extra PQ columns. Leave 0 if not used.",
        )
        st.divider()
        section_form(
            "hoa_expense",
            "Estate / HOA paid to the estate",
            "What we pay Xanadu (or any master HOA). Not included in ordinary levies.",
        )
        st.divider()
        section_form("other", "Other Income", "Fixed Eskom / rental / interest live here. Leave unused lines at 0.")
        st.divider()
        section_form("recoveries_other", "Other recoveries", "Insurance / legal recoveries. Utility recoveries sit under Municipal.")
        st.divider()
        section_form("fixed", "Fixed monthly charges on the owner invoice", "Optional. Mount Kos: insurance + prepaid + Eskom fixed.")

    with tabs[2]:
        section_form("municipal", "Municipal charges", "Gross amount on its own line. Recovery on the ‘Less:’ line.")

    with tabs[3]:
        section_form("expenditure", "Expenditure", "Operating costs except R&M, personnel and tax. Add or delete rows as needed.")

    with tabs[4]:
        section_form(
            "rm",
            "Repair and Maintenance",
            "Type an insurance payout on the line it belongs to, then Save. Net = yearly − payout.",
            rm=True,
        )

    with tabs[5]:
        section_form("personnel", "Personnel", "Salaries, casuals, PAYE/UIF, bonuses.")

    with tabs[6]:
        section_form("tax", "Income Tax", "Most packs forget this. If the financial statements show tax, budget it here.")

    with tabs[7]:
        section_form("special", "Special Projects", "Year 1 of the 10-year plan. Tick the sidebar box only if levies must fund it.")

    with tabs[8]:
        st.subheader("PQ / levy schedule")
        st.caption("Same columns as Thornhill: Levies, estate/HOA lines, insurance (if billed), reserve, CSOS. Each unit = PQ × that column’s monthly total.")
        pq_file = st.file_uploader("PQ Excel or CSV", type=["csv", "xlsx", "xls"], key="pqfile")
        if pq_file:
            raw = pd.read_csv(pq_file) if pq_file.name.lower().endswith(".csv") else pd.read_excel(pq_file)
            raw.columns = [str(c).strip() for c in raw.columns]
            seen, cols = {}, []
            for c in raw.columns:
                if c in seen:
                    seen[c] += 1
                    cols.append(f"{c}_{seen[c]}")
                else:
                    seen[c] = 0
                    cols.append(c)
            raw.columns = cols
            unit_col = next((c for c in raw.columns if re.search(r"unit|owner|code", str(c), re.I)), None)
            pq_col = next((c for c in raw.columns if re.search(r"pq|quota|ratio|^share$", str(c), re.I)), None)
            if not unit_col or not pq_col:
                st.error("Need a Unit column and a PQ column. Found: " + ", ".join(map(str, raw.columns)))
            else:
                clean = pd.DataFrame({
                    "Unit": raw[unit_col].astype(str).str.strip(),
                    "PQ": pd.to_numeric(raw[pq_col], errors="coerce").fillna(0),
                })
                clean = clean[clean["Unit"].str.lower().ne("nan") & clean["Unit"].ne("")].reset_index(drop=True)
                total = clean["PQ"].sum()
                if 50 < total < 150:
                    clean["PQ"] = clean["PQ"] / 100
                    st.info(f"PQ looked like percentages (total {total:.2f}). Divided by 100.")
                st.session_state.pq = clean.to_dict("records")
                st.success(f"{len(clean)} units. PQ total {clean['PQ'].sum():.6f}")
        if st.session_state.pq:
            prev = pd.DataFrame(st.session_state.pq)
            bills = pq_bill_lines(st.session_state)
            extra_cols = []
            for name, yearly in bills:
                prev[name] = prev["PQ"] * (yearly / 12)
                extra_cols.append(name)
            if extra_cols:
                prev["Total monthly"] = prev[extra_cols].sum(axis=1)
            st.dataframe(prev, use_container_width=True, hide_index=True)
            st.caption("Monthly column totals: " + " · ".join(f"{n} {money(y/12)}" for n, y in bills))

    with tabs[9]:
        st.subheader("10-year maintenance plan")
        st.caption("Paste from Excel (name + 10 year amounts) or type in the table and Save.")
        paste = st.text_area("Paste from Excel", height=120)
        if st.button("Paste into plan") and paste.strip():
            parsed = []
            for line in paste.splitlines():
                line = line.strip()
                if not line or re.match(r"^(project|description)", line, re.I):
                    continue
                parts = [p.strip() for p in re.split(r"\t|;|,|\s{2,}", line) if p.strip()]
                if not parts:
                    continue
                years, yi = [0.0] * 10, 0
                for p in parts[1:]:
                    if yi >= 10:
                        break
                    n = num(p)
                    if n is None:
                        continue
                    years[yi] = n
                    yi += 1
                parsed.append({"desc": parts[0], "years": years})
            if parsed:
                st.session_state.ymp = parsed
                st.success(f"Loaded {len(parsed)} projects.")
        ymp_rows = []
        for p in st.session_state.ymp:
            rec = {"Project": p.get("desc") or ""}
            years = p.get("years") or [0] * 10
            for i in range(10):
                rec[f"Y{i+1}"] = float(years[i] if i < len(years) else 0)
            ymp_rows.append(rec)
        with st.form("ymp_form"):
            ed = st.data_editor(pd.DataFrame(ymp_rows), num_rows="dynamic", use_container_width=True, hide_index=True)
            if st.form_submit_button("Save 10-year plan"):
                st.session_state.ymp = [{
                    "desc": str(r.get("Project") or ""),
                    "years": [float(r.get(f"Y{i+1}") or 0) for i in range(10)],
                } for _, r in ed.iterrows()]
                st.success("Saved.")
        if st.button("Copy Year 1 into Special Projects"):
            spec = []
            for p in st.session_state.ymp:
                y1 = float((p.get("years") or [0])[0] or 0)
                if y1:
                    rec = row(p["desc"], "From 10-year plan Year 1")
                    rec["yearly"] = y1
                    spec.append(rec)
            if spec:
                st.session_state.sections["special"] = spec
                st.success(f"Copied {len(spec)} projects.")

    with tabs[10]:
        st.subheader("Download Excel")
        st.caption("Budget + PQ + 10-year plan. Yellow cells are inputs.")
        if not st.session_state.complex_name:
            st.warning("Type the complex name in the sidebar first.")
        elif st.button("Build Excel file", type="primary"):
            xls = generate_excel(st.session_state)
            st.session_state["xlsx"] = xls.getvalue()
        if st.session_state.get("xlsx"):
            name = re.sub(r"\s+", "_", st.session_state.complex_name)
            st.download_button(
                "Download budget Excel",
                data=st.session_state["xlsx"],
                file_name=f"Budget_{name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
